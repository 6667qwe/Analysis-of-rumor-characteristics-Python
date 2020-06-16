import requests
import json
from openpyxl import Workbook
import os
import datetime

Title = []  # 存储新闻标题
LinkUrl = []  # 存储新闻的链接
keyword = []  # 存储新闻性质
SourceName = []  # 存储新闻来源
PubTime = []  # 存储新闻发布时间
# L_url = []  # 存储新闻的完整链接
header_url = "http://www.piyao.org.cn"


def getHTMLMassages():
    headers = {
        'user-agent': 'Mozilla/5.0'
    }
    for i in range(85):
        i = i + 1
        url = 'http://qc.wa.news.cn/nodeart/list?nid=11215616&pgnum=' + str(
            i) + '&cnt=10&attr=63&tp=1&orderby=1'
        # 爬取页面
        try:
            r = requests.get(url, headers=headers, timeout=30)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            # print(r.text[:1000])
            # print(r.status_code)
        except:
            print(("爬取页面异常"))
            print(r.status_code)
        start_results = r.text
        start_Loc = start_results.find('{')
        results = start_results[start_Loc:-2]
        results = json.loads(results)

        # 遍历一页的新闻数（子循环1号开始）
        for j in range(len(results['data']['list'])):
            print(j)
            # 添加新数据
            if results['data']['list'][j]['Title'] not in Title:
                Title.append(results['data']['list'][j]['Title'])  # 获取新闻标题
                LinkUrl.append(results['data']['list'][j]['LinkUrl'])  # 获取新闻链接
                keyword.append(results['data']['list'][j]['keyword'])  # 获取新闻性质
                SourceName.append(results['data']['list'][j]['SourceName'])  # 获取新闻来源
                PubTime.append(results['data']['list'][j]['PubTime'])  # 获取新闻发布时间
        # 遍历一页的新闻数（子循环1号结束）

        for index in range(len(Title)):
            print('标题：', Title[index], LinkUrl[index], keyword[index], SourceName[index], PubTime[index])
            # 给链接加上"http://www.piyao.org.cn"，存储新闻的完整链接
            '''
            if 'http' not in LinkUrl[index]:
                L_url.append(header_url + LinkUrl[index])
                # print('新闻链接：', header_url + LinkUrl[index])
            else:
                # print('新闻链接：', LinkUrl[index])
                L_url.append(LinkUrl[index])
            '''
            # print('源链接：', url+source_url[index])
            # print(len(Title))  # 获取的新闻数量
        # 遍历一页的新闻数（子循环2号结束）
    # （总循环结束）


def saveData():
    # 存储数据到xlxs文件,即excel表格
    wb = Workbook()
    if not os.path.exists(root):  # 判断文件夹是否存在
        os.mkdir(root)  # 新建存储文件夹
    filename = root + '疫情辟谣' + datetime.datetime.now().strftime(
        '%Y-%m-%d-%H-%m') + '.xlsx'  # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'  # 更改工作表的标题
    ws['A1'] = '标题'  # 对表格加入标题
    ws['B1'] = '新闻链接'
    ws['C1'] = '新闻性质'
    ws['D1'] = '新闻来源'
    ws['E1'] = '新闻发布时间'
    for row in range(2, len(Title) + 2):  # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=Title[row - 2])
        _ = ws.cell(column=2, row=row, value=LinkUrl[row - 2])
        _ = ws.cell(column=3, row=row, value=keyword[row - 2])
        _ = ws.cell(column=4, row=row, value=SourceName[row - 2])
        _ = ws.cell(column=5, row=row, value=PubTime[row - 2])
    wb.save(filename=filename)  # 保存文件
    print("文件已保存")


if __name__ == '__main__':
    root = 'C:/Users/赵宗天/Desktop/认识Python/爬取文档/'
    getHTMLMassages()
    saveData()
