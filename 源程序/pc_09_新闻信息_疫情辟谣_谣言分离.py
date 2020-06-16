import xlrd
from openpyxl import Workbook

path = 'C://Users//赵宗天//Desktop//认识Python//爬取文档//'
file = '疫情辟谣2020-05-13-00-05.xlsx'
oldfile = path + file

Title = []  # 存储新闻标题
LinkUrl = []  # 存储新闻的链接
keyword = []  # 存储新闻性质
SourceName = []  # 存储新闻来源
PubTime = []  # 存储新闻发布时间

Title2 = []  # 存储新闻标题
LinkUrl2 = []  # 存储新闻的链接
keyword2 = []  # 存储新闻性质
SourceName2 = []  # 存储新闻来源
PubTime2 = []  # 存储新闻发布时间

new_file1 = path + '疫情辟谣2020-06-02-00-05_谣言+辟谣.xlsx'
new_file2 = path + '疫情辟谣2020-06-02-00-05_事实.xlsx'


def separateRumour():
    # 打开excel文件，创建一个workbook对象,book对象也就是fruits.xlsx文件,表含有sheet名
    rbook = xlrd.open_workbook(oldfile)
    # sheets方法返回对象列表,[<xlrd.sheet.Sheet object at 0x103f147f0>]
    rbook.sheets()
    # xls默认有3个工作簿,Sheet1,Sheet2,Sheet3
    rsheet = rbook.sheet_by_index(0)  # 取第一个工作簿
    # 循环工作簿的所有行
    for row in rsheet.get_rows():
        keyword_column = row[2]  # 品名所在的列
        keyword_value = keyword_column.value  # 项目名
        Title_column = row[0]
        Title_value = Title_column.value
        LinkUrl_column = row[1]
        LinkUrl_value = LinkUrl_column.value
        SourceName_column = row[3]
        SourceName_value = SourceName_column.value
        PubTime_column = row[4]
        PubTime_value = PubTime_column.value
        if keyword_value == '谣言' or keyword_value == '辟谣':  # 排除第一行
            Title.append(Title_value)
            LinkUrl.append(LinkUrl_value)
            keyword.append(keyword_value)
            SourceName.append(SourceName_value)
            PubTime.append(PubTime_value)
        elif keyword_value == '事实':
            Title2.append(Title_value)
            LinkUrl2.append(LinkUrl_value)
            keyword2.append(keyword_value)
            SourceName2.append(SourceName_value)
            PubTime2.append(PubTime_value)


def writeText():
    wb = Workbook()
    wb2 = Workbook()
    ws = wb.active
    ws.title = 'rumour'  # 更改工作表的标题
    ws['A1'] = '标题'  # 对表格加入标题
    ws['B1'] = '新闻链接'
    ws['C1'] = '新闻性质'
    ws['D1'] = '新闻来源'
    ws['E1'] = '新闻发布时间'
    for row in range(2, len(Title) + 2):  # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=Title[row - 2])
        _ = ws.cell(column=2, row=row, value=LinkUrl[row - 2])
        _ = ws.cell(column=3, row=row, value=keyword[row - 2])
        _ = ws.cell(column=4, row=row, value=SourceName[row - 2])
        _ = ws.cell(column=5, row=row, value=PubTime[row - 2])
    wb.save(filename=new_file1)  # 保存文件

    ws2 = wb2.active
    ws2.title = 'factor'  # 更改工作表的标题
    ws2['A1'] = '标题'  # 对表格加入标题
    ws2['B1'] = '新闻链接'
    ws2['C1'] = '新闻性质'
    ws2['D1'] = '新闻来源'
    ws2['E1'] = '新闻发布时间'
    for row in range(2, len(Title2) + 2):  # 将数据写入表格
        _ = ws2.cell(column=1, row=row, value=Title2[row - 2])
        _ = ws2.cell(column=2, row=row, value=LinkUrl2[row - 2])
        _ = ws2.cell(column=3, row=row, value=keyword2[row - 2])
        _ = ws2.cell(column=4, row=row, value=SourceName2[row - 2])
        _ = ws2.cell(column=5, row=row, value=PubTime2[row - 2])
    wb2.save(filename=new_file2)  # 保存文件
    print('保存成功')


if __name__ == '__main__':
    separateRumour()
    writeText()
